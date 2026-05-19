"""巡检中心 API。"""
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.api.deps import get_current_api_user, api_permission_required, get_client_ip
from app.core.settings import get_config, get_config_float, set_config
from app.db.database import get_db
from app.models.user import User
from app.services.audit import write_log
from app.services.patrol import (
    delete_report,
    get_report,
    get_report_items,
    list_reports,
    run_patrol,
)

router = APIRouter(prefix="/patrol", tags=["巡检中心"])

# ─── 阈值配置 ───────────────────────────────────────────────

_PATROL_THRESHOLD_KEYS = {
    "patrol.cpu_warning": "CPU 使用率警告阈值（%）",
    "patrol.cpu_critical": "CPU 使用率严重阈值（%）",
    "patrol.memory_warning": "内存使用率警告阈值（%）",
    "patrol.memory_critical": "内存使用率严重阈值（%）",
    "patrol.disk_warning": "磁盘使用率警告阈值（%）",
    "patrol.disk_critical": "磁盘使用率严重阈值（%）",
    "patrol.load_warning": "系统负载警告阈值",
    "patrol.load_critical": "系统负载严重阈值",
}


class ThresholdUpdate(BaseModel):
    value: str


@router.get("/thresholds")
def api_get_thresholds(
    db: Session = Depends(get_db),
    _: User = Depends(api_permission_required("patrol.view")),
):
    """获取巡检阈值配置。"""
    items = []
    for key, desc in _PATROL_THRESHOLD_KEYS.items():
        items.append({
            "key": key,
            "value": get_config(db, key),
            "description": desc,
        })
    return {"code": 0, "data": {"items": items}}


@router.put("/thresholds/{key}")
def api_update_threshold(
    key: str,
    body: ThresholdUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(api_permission_required("patrol.execute")),
):
    """更新单个巡检阈值。"""
    if key not in _PATROL_THRESHOLD_KEYS:
        raise HTTPException(status_code=400, detail=f"不支持的配置项: {key}")
    # 校验是合法数字
    try:
        float(body.value)
    except ValueError:
        raise HTTPException(status_code=400, detail="阈值必须为数字")
    set_config(db, key, body.value.strip(), _PATROL_THRESHOLD_KEYS[key])
    write_log(db, user=current_user, action="update", target_type="settings",
              target_name=key, detail=f"巡检阈值更新为 {body.value.strip()}",
              ip_address=get_client_ip(request))
    db.commit()
    return {"code": 0, "msg": "阈值已更新"}


@router.put("/thresholds")
def api_update_thresholds_bulk(
    body: dict[str, str],
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(api_permission_required("patrol.execute")),
):
    """批量更新巡检阈值。"""
    updated = []
    for key, value in body.items():
        if key not in _PATROL_THRESHOLD_KEYS:
            continue
        try:
            float(value)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"{key} 阈值必须为数字")
        set_config(db, key, value.strip(), _PATROL_THRESHOLD_KEYS[key])
        updated.append(key)
    if updated:
        write_log(db, user=current_user, action="update", target_type="settings",
                  target_name="patrol.thresholds", detail=f"批量更新 {len(updated)} 项巡检阈值",
                  ip_address=get_client_ip(request))
        db.commit()
    return {"code": 0, "msg": f"已更新 {len(updated)} 项阈值"}


@router.post("/run")
async def api_run_patrol(
    db: Session = Depends(get_db),
    current_user: User = Depends(api_permission_required("patrol.execute")),
):
    """手动触发一次巡检。"""
    report = await run_patrol(db, operator=current_user.full_name or current_user.username)
    return {
        "code": 0,
        "msg": "巡检完成",
        "data": {
            "id": report.id,
            "title": report.title,
            "status": report.status,
            "total_checks": report.total_checks,
            "normal_count": report.normal_count,
            "warning_count": report.warning_count,
            "critical_count": report.critical_count,
            "summary": report.summary,
            "created_at": report.created_at.isoformat(),
        },
    }


@router.get("/reports")
def api_list_reports(
    status: str = "",
    page: int = 1,
    page_size: int = 20,
    db: Session = Depends(get_db),
    _: User = Depends(api_permission_required("patrol.view")),
):
    """查询巡检报告列表。"""
    offset = (max(page, 1) - 1) * page_size
    items, total = list_reports(db, status=status, limit=page_size, offset=offset)
    return {
        "code": 0,
        "data": {
            "items": [
                {
                    "id": r.id,
                    "title": r.title,
                    "status": r.status,
                    "total_checks": r.total_checks,
                    "normal_count": r.normal_count,
                    "warning_count": r.warning_count,
                    "critical_count": r.critical_count,
                    "summary": r.summary,
                    "operator": r.operator,
                    "created_at": r.created_at.isoformat(),
                }
                for r in items
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    }


@router.get("/reports/{report_id}")
def api_report_detail(
    report_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(api_permission_required("patrol.view")),
):
    """获取巡检报告详情（含所有巡检项）。"""
    report = get_report(db, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="报告不存在")

    items = get_report_items(db, report_id)

    return {
        "code": 0,
        "data": {
            "report": {
                "id": report.id,
                "title": report.title,
                "status": report.status,
                "total_checks": report.total_checks,
                "normal_count": report.normal_count,
                "warning_count": report.warning_count,
                "critical_count": report.critical_count,
                "summary": report.summary,
                "operator": report.operator,
                "created_at": report.created_at.isoformat(),
            },
            "items": [
                {
                    "id": item.id,
                    "category": item.category,
                    "target_name": item.target_name,
                    "target_ip": item.target_ip,
                    "check_name": item.check_name,
                    "status": item.status,
                    "value": item.value,
                    "threshold": item.threshold,
                    "detail": item.detail,
                }
                for item in items
            ],
        },
    }


@router.get("/reports/{report_id}/export")
def api_export_report(
    report_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(api_permission_required("patrol.view")),
):
    """导出巡检报告为 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    report = get_report(db, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="报告不存在")

    items = get_report_items(db, report_id)

    # ── 创建 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "巡检报告"

    # 样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    status_fills = {
        "normal": PatternFill("solid", fgColor="C6EFCE"),
        "warning": PatternFill("solid", fgColor="FFEB9C"),
        "critical": PatternFill("solid", fgColor="FFC7CE"),
    }
    status_labels = {"normal": "正常", "warning": "警告", "critical": "严重"}
    category_labels = {"host": "主机巡检", "k8s": "K8s 巡检", "asset": "资产巡检"}

    # 标题行
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = report.title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 报告信息
    info_rows = [
        ("操作人", report.operator or "-", "巡检时间", report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "-"),
        ("总检查项", str(report.total_checks), "摘要", report.summary),
        ("正常", str(report.normal_count), "警告", str(report.warning_count)),
    ]
    for i, (k1, v1, k2, v2) in enumerate(info_rows, start=3):
        ws.cell(i, 1, k1).font = Font(bold=True)
        ws.cell(i, 2, v1)
        ws.cell(i, 4, k2).font = Font(bold=True)
        ws.cell(i, 5, v2)

    # 表头
    headers = ["分类", "目标名称", "IP", "检查项", "状态", "当前值", "阈值", "详情"]
    header_row = 7
    for col, h in enumerate(headers, 1):
        cell = ws.cell(header_row, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行
    for i, item in enumerate(items, start=header_row + 1):
        row_data = [
            category_labels.get(item.category, item.category),
            item.target_name,
            item.target_ip,
            item.check_name,
            status_labels.get(item.status, item.status),
            item.value,
            item.threshold,
            item.detail,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(i, col, val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
        # 状态列着色
        fill = status_fills.get(item.status)
        if fill:
            ws.cell(i, 5).fill = fill

    # 列宽
    col_widths = [12, 20, 16, 16, 10, 16, 28, 32]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 输出
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{report.title}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


@router.delete("/reports/{report_id}")
def api_delete_report(
    report_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(api_permission_required("patrol.delete")),
):
    """删除巡检报告。"""
    report = get_report(db, report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="报告不存在")
    delete_report(db, report)
    return {"code": 0, "msg": "删除成功"}
